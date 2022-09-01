from PIL import Image,ImageDraw,ImageFont
import openpyxl
from PIL import Image, ImageDraw, ImageFont

wb = openpyxl.load_workbook('C:/ImgP/Details/moviedetails.xlsx') #excel location
sh = wb.active
n = 2 #number of organizers 


for i in range(n):
    background = Image.open('C:/ImgP/Images/Background.png')
    
    draw = ImageDraw.Draw(background)
    newfont = ImageFont.truetype('C:/Windows/Fonts/ARLRDBD.TTF', size = 40)
    poster_width = background.width
    
    c = sh.cell(row=i+2,column=3)
    director = str(c.value)
    b = sh.cell(row=i+2,column=4)
    producer = str(b.value)
    d=sh.cell(row=i+2,column=5)
    writer= str(d.value)
    a= sh.cell(row=i+2,column=6)
    cast= str(a.value)
    e= sh.cell(row=i+2,column=7)
    music = str(e.value)
    f=sh.cell(row=i+2,column=2)
    name_= str(f.value)
    name=name_.replace(" ","_")
    
    movie=Image.open('C:/ImgP/Images/'+name+'.png')
    movie.thumbnail((1000,1500))
    background.paste(movie,(int((poster_width-movie.width)/2),100))
    text = "DIRECTOR : "+director
    text_width, _ = draw.textsize(text, font=newfont)
    draw.text((200,1660),text, font = newfont, fill=(0,0,0))
    text = "PRODUCER : " + producer
    text_width, _ = draw.textsize(text, font=newfont)
    draw.text((200,1700),text, font = newfont, fill=(0,0,0))
    text = "CAST : "+ cast
    text_width, _ = draw.textsize(text, font=newfont)
    draw.text((200,1740),text, font = newfont, fill=(0,0,0))
    text = "MUSIC : "+ music
    text_width, _ = draw.textsize(text, font=newfont)
    draw.text((200,1780),text, font = newfont, fill=(0,0,0))
    text = "WRITER : "+ writer
    text_width, _ = draw.textsize(text, font=newfont)
    draw.text((200,1820),text, font = newfont, fill=(0,0,0))
    im_1 = background.convert('RGB')
    im_1.save('C:/ImgP/finalposter/'+name+'.png') 
    im_1.show()